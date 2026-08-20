from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

STOCK_HEADER_HINTS = (
    "trading_symbol",
    "tradingsymbol",
    "nse_symbol",
    "symbol",
    "qty",
    "quantity",
    "average_price",
    "avg_price",
    "avg_buy",
    "ltp",
)
MF_HEADER_HINTS = (
    "scheme",
    "scheme_name",
    "folio",
    "units",
    "nav",
    "amc",
    "sub-category",
    "sub_category",
)


@dataclass
class ImportResult:
    kind: str  # stocks | mf | mixed
    stocks: list[dict[str, Any]] = field(default_factory=list)
    mutual_funds: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _norm_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (value or "").strip().lower()).strip("_")


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    text = text.replace(",", "").replace("₹", "").replace("%", "")
    text = text.replace("(", "-").replace(")", "")
    if not text or text in ("-", "na", "n/a", "none"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _read_tabular(path: Path) -> tuple[list[str], list[list[Any]]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        raw = path.read_text(encoding="utf-8-sig", errors="replace")
        sample = raw[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(io.StringIO(raw), dialect))
        if not rows:
            return [], []
        return [str(c) for c in rows[0]], rows[1:]

    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return [], []
        headers = ["" if c is None else str(c) for c in header_row]
        data = [list(r) for r in rows_iter]
        wb.close()
        return headers, data

    if suffix == ".numbers":
        return _read_numbers(path)

    raise ValueError(f"Unsupported file type: {suffix}. Use CSV, XLSX, or Numbers.")


def _read_numbers(path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        from numbers_parser import Document
    except ImportError as exc:
        raise ValueError(
            "Numbers files need numbers-parser. Convert to CSV/XLSX in Numbers, or: pip install numbers-parser"
        ) from exc

    doc = Document(str(path))
    best_headers: list[str] = []
    best_rows: list[list[Any]] = []
    for sheet in doc.sheets:
        for table in sheet.tables:
            header_idx = None
            for r in range(min(table.num_rows, 40)):
                values = [table.cell(r, c).value for c in range(table.num_cols)]
                norms = [_norm_header(str(v or "")) for v in values]
                if _detect_kind(norms) != "unknown":
                    header_idx = r
                    best_headers = ["" if v is None else str(v) for v in values]
                    break
            if header_idx is None:
                continue
            data = []
            for r in range(header_idx + 1, table.num_rows):
                data.append([table.cell(r, c).value for c in range(table.num_cols)])
            if len(data) > len(best_rows):
                best_rows = data
    if not best_headers:
        raise ValueError("Could not find a holdings table in the Numbers file.")
    return best_headers, best_rows


def _detect_kind(norm_headers: list[str]) -> str:
    joined = " ".join(norm_headers)
    mf_score = sum(1 for h in MF_HEADER_HINTS if h in joined or h in norm_headers)
    stock_score = sum(1 for h in STOCK_HEADER_HINTS if h in joined or h in norm_headers)
    if "scheme_name" in norm_headers or "scheme" in norm_headers or "folio_no" in norm_headers:
        return "mf"
    if "trading_symbol" in norm_headers or "isin" in norm_headers:
        if "units" in norm_headers and "quantity" not in norm_headers:
            return "mf"
        return "stocks"
    if mf_score > stock_score and mf_score >= 2:
        return "mf"
    if stock_score >= 2:
        return "stocks"
    return "unknown"


def _first(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def _parse_stock_rows(mapped: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_symbol: dict[str, dict[str, Any]] = {}
    for row in mapped:
        symbol = str(
            _first(row, "trading_symbol", "tradingsymbol", "nse_symbol", "symbol", "stock_name", "scrip")
            or ""
        ).strip().upper()
        symbol = re.sub(r"\s+", "", symbol)
        symbol = symbol.replace(".NS", "").replace("-EQ", "")
        if not symbol or symbol in {"SCHEME", "NAME", "TOTAL"}:
            continue
        qty = _to_float(_first(row, "quantity", "qty", "net_qty", "net_quantity", "shares"))
        invested = _to_float(
            _first(row, "invested_value", "invested", "invested_amount", "buy_value", "total_invested")
        )
        avg = _to_float(
            _first(
                row,
                "average_price",
                "avg_price",
                "avg_buy_price",
                "average_buy_price",
                "avg_buy",
                "buy_price",
            )
        )
        if qty is None or qty <= 0:
            continue
        if avg is None and invested and qty:
            avg = invested / qty
        if avg is None:
            continue
        if invested is None:
            invested = qty * avg
        existing = by_symbol.get(symbol)
        if existing:
            new_qty = existing["quantity"] + qty
            new_inv = existing["invested_value"] + invested
            existing["quantity"] = new_qty
            existing["invested_value"] = new_inv
            existing["average_price"] = new_inv / new_qty if new_qty else avg
        else:
            by_symbol[symbol] = {
                "trading_symbol": symbol,
                "exchange": "NSE",
                "quantity": qty,
                "average_price": avg,
                "invested_value": invested,
            }
    return list(by_symbol.values())


def _parse_mf_rows(mapped: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in mapped:
        name = str(_first(row, "scheme_name", "scheme", "fund_name", "name") or "").strip()
        if not name or name.lower() in {"scheme name", "total", "holdings as on"}:
            continue
        units = _to_float(_first(row, "units", "unit", "quantity"))
        invested = _to_float(_first(row, "invested_value", "invested", "invested_amount", "amount"))
        avg_nav = _to_float(_first(row, "avg_nav", "average_nav", "avg_price", "purchase_nav"))
        category = str(_first(row, "sub_category", "subcategory", "category") or "").strip()
        if units is None or units <= 0:
            continue
        if avg_nav is None and invested and units:
            avg_nav = invested / units
        if avg_nav is None:
            continue
        if invested is None:
            invested = units * avg_nav
        key = name.lower()
        existing = grouped.get(key)
        if existing:
            new_units = existing["units"] + units
            new_inv = existing["invested_value"] + invested
            existing["units"] = new_units
            existing["invested_value"] = new_inv
            existing["avg_nav"] = new_inv / new_units if new_units else avg_nav
        else:
            grouped[key] = {
                "name": name,
                "units": units,
                "avg_nav": avg_nav,
                "invested_value": invested,
                "category": category,
            }
    return list(grouped.values())


def resolve_mf_scheme_codes(holdings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    from grow_trade_assistant.mf_config import search_schemes

    resolved: list[dict[str, Any]] = []
    for h in holdings:
        query = h["name"]
        results = search_schemes(query, limit=8)
        growth = [
            r
            for r in results
            if "growth" in r["name"].lower() and "idcw" not in r["name"].lower() and "bonus" not in r["name"].lower()
        ]
        preferred = growth or results
        direct = [r for r in preferred if "direct" in r["name"].lower()]
        pick = (direct or preferred)[:1]
        if pick:
            h = dict(h)
            h["scheme_code"] = pick[0]["scheme_code"]
            h["name"] = pick[0]["name"]
        else:
            logger.warning("Could not resolve scheme code for %s", query)
        resolved.append(h)
    return resolved


def parse_groww_file(path: Path, kind: str = "auto") -> ImportResult:
    headers, rows = _read_tabular(path)
    # Groww reports often have title rows before the real header
    norm = [_norm_header(h) for h in headers]
    detected = _detect_kind(norm)
    if detected == "unknown":
        for i, row in enumerate(rows[:40]):
            candidate = ["" if c is None else str(c) for c in row]
            cand_norm = [_norm_header(c) for c in candidate]
            cand_kind = _detect_kind(cand_norm)
            if cand_kind != "unknown":
                headers = candidate
                rows = rows[i + 1 :]
                norm = cand_norm
                detected = cand_kind
                break
    if kind == "auto":
        kind = detected if detected != "unknown" else "stocks"
    mapped = []
    for row in rows:
        if not any(cell not in (None, "") for cell in row):
            continue
        item: dict[str, Any] = {}
        for i, key in enumerate(norm):
            if not key:
                continue
            item[key] = row[i] if i < len(row) else None
        mapped.append(item)

    result = ImportResult(kind=kind)
    if kind in ("stocks", "mixed"):
        result.stocks = _parse_stock_rows(mapped)
    if kind in ("mf", "mixed"):
        result.mutual_funds = _parse_mf_rows(mapped)
        result.mutual_funds = resolve_mf_scheme_codes(result.mutual_funds)
    if not result.stocks and not result.mutual_funds:
        result.warnings.append("No holdings parsed. Check that the file is a Groww stocks or MF export.")
    return result


def save_stocks_file(path: Path, holdings: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(holdings, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def load_stocks_file(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        return data.get("holdings", [])
    return data if isinstance(data, list) else []
